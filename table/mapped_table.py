import types
from typing import Iterable, Optional, Union, List, Sequence
from functools import partial
from openpyxl import load_workbook
from .mapped_sequence import MappedSequence
from .aggregation import mean_aggregate
from .formatter import HtmlFormatter
from .utils import is_iterable, is_scalar


class MappedTable:
    """A generic container for immutable 2-dimensional data"""

    __slots__ = ['_column_values', '_row_values', '_index', '_columns']

    def __init__(self, values: Sequence[Sequence], columns: Sequence[str], index: Optional[Iterable] = None,
                 axis=0):
        """

        Parameters
        ----------
        values: Iterable[Iterable]
            Typically 2D-nested list of rows
        columns: Iterable
            List of the names of the columns
        index : Optional[Iterable]
            List of the names of the rows
        axis: int
            Orientation of the values. If axis = 0, values are Iterable of rows. If axis = 1, values are Iterable of
            columns.
        """
        # index and columns are also used as keys to ease slicing.
        self._columns = MappedSequence(columns, columns)

        if axis == 0:
            if index is None:
                index = range(len(values))
            row_values = [MappedSequence(value, keys=columns, name=idx) for idx, value in zip(index, values)]
            row_values = MappedSequence(row_values, keys=index)
            column_values = None

        else:
            if index is None:
                index = range(len(values[0]))
            column_values = [MappedSequence(value, keys=index, name=col) for col, value in zip(columns, values)]
            column_values = MappedSequence(column_values, keys=columns)
            row_values = None

        self._index = MappedSequence(index, index)

        # Store as MappedSequence of columns
        self._column_values = column_values
        # Store as MappedSequence of rows
        self._row_values = row_values

    @classmethod
    def from_excel(cls, file_path, header: Optional[Union[int, Iterable[int]]] = 0,
                   sheetname: Optional[str] = None, skiprows: Optional[int] = None):
        """

        Parameters
        ----------
        file_path: path to excel file
        header: Optional[Union[int, Iterable[int]]]
            If header is None, the columns are defined as a sequence of integers.
        sheetname: Optional[str]
            Name of the sheet to parse. If None, the first will used instead.
        skiprows: Optional[int]

        Returns
        -------

        """
        # Get workbook
        workbook = load_workbook(file_path, data_only=True)
        # Handle sheetname
        # By default, parse the first sheet in the excel
        if sheetname is None:
            sheetname = workbook.sheetnames[0]
        sheet = workbook[sheetname]

        # Read all rows of the sheet

        if skiprows is None:
            skiprows = 0
        else:
            skiprows += 1
        values = [row for n, row in enumerate(sheet.iter_rows(values_only=True, min_row=skiprows))]

        # Handle header
        # If none, the header is simply a numeric index
        if header is None:
            columns = range(len(values[0]))

        elif is_iterable(header):
            # header is a list of tuple
            columns = list(
                zip(*[row for row in sheet.iter_rows(min_row=min(header) + skiprows, max_row=max(header) + 1 + skiprows,
                                                     values_only=True)])
            )
            values = values[max(header):]

        elif type(header) == int:
            columns = [row for row in
                       sheet.iter_rows(min_row=header + skiprows, max_row=header + 1 + skiprows, values_only=True)][0]
            values = values[header + 1:]

        else:
            raise ValueError

        return cls(values=values, columns=columns, axis=0)

    @property
    def index(self):
        return self._index

    @property
    def columns(self):
        return self._columns

    @property
    def values(self) -> 'MappedSequence':
        return self.column_values

    @property
    def row_values(self) -> 'MappedSequence[MappedSequence]':
        if self._row_values is None:
            # if row values is None, then _column_values holds the data
            row_values = [MappedSequence(value, keys=self.columns, name=idx) for idx, *value in
                          zip(self.index, *self._column_values)]
            self._row_values = MappedSequence(row_values, keys=self.index)
        return self._row_values

    @property
    def column_values(self):
        if self._column_values is None:
            column_values = [MappedSequence(value, keys=self.index, name=col) for col, *value in
                             zip(self.columns, *self._row_values)]
            self._column_values = MappedSequence(column_values, keys=self.columns)
        return self._column_values

    @property
    def T(self):
        return self.transpose()

    def transpose(self):
        return self.row_values

    def __str__(self):
        """
        Print an ascii sample of the contents of this sequence.
        """

        return self.values.__str__()

    def __repr__(self):
        return self.__str__()

    def _repr_html_(self) -> str:
        formatter = HtmlFormatter()
        return formatter.format_table(self)

    def __getitem__(self, item):
        # 1-dimensional item
        if type(item) is slice:
            # Slice the rows
            new_index = self.index[item]
            new_values = [value[item] for value in self.values]
            if isinstance(new_index, MappedSequence):
                return MappedTable(values=new_values, index=new_index, columns=self.columns, axis=1)
            else:
                return MappedSequence(values=new_values, keys=self.columns, name=new_index)
        elif type(item) is list:
            # Get the corresponding columns
            values = self.values[item]  # getting MappedSequence
            return MappedTable(values=values, columns=values.keys(), index=self.index, axis=1)

        elif type(item) is tuple and len(item) == 2:
            new_columns = self.columns[item[1]]
            new_index = self.index[item[0]]
            subset = self.values[item[1]]
            sequence_colums = isinstance(new_columns, MappedSequence)
            sequence_index = isinstance(new_index, MappedSequence)
            # When slicing multiple columns and multiple rows
            if sequence_colums and sequence_index:
                new_values = [value[item[0]] for value in subset]
                return MappedTable(values=new_values, columns=new_columns, index=new_index, axis=1)
            # When slicing multiple columns but one rows
            elif sequence_colums and not sequence_index:
                new_values = [value[item[0]] for value in subset]
                return MappedSequence(new_values, keys=new_columns, name=new_index)
            elif not sequence_colums and sequence_index:
                # If the new index is a MappedSequence, should return a MappedSequence
                new_values = subset[item[0]]
                return MappedSequence(values=new_values, keys=new_index, name=new_columns)
            elif not sequence_colums and not sequence_index:
                # If the new index is only a scalar, should return a scalar
                return subset[item[0]]
            else:
                raise KeyError

        else:
            return self.values[item]

    def __getattr__(self, k):
        if k in self.columns:
            return self[k]
        else:
            return self.__getattribute__(k)

    def __len__(self):
        return len(self.index)

    def __eq__(self, other: 'MappedTable') -> bool:
        result = self.columns == other.columns
        result &= self.values == other.values
        return result

    @property
    def shape(self):
        return len(self.index), len(self.columns)

    def sort_values(self, key: Union[int, str, Iterable[str]], ascending: bool = True) -> 'MappedTable':
        """

        Parameters
        ----------
        key: Union[int, str, Iterable[str]]
            key used to sort the table. It can be an integer (position of the column) or the name of the column
            passed as a string or a list of string.
        ascending : bool
            sort values in ascending ordre if True

        Returns
        -------
        MappedTable
        Sorted table
        """
        values_as_rows = self.to_list()
        sort_key = partial(MappedSequence.__getitem__, item=key)
        sorted_values = sorted(values_as_rows, key=sort_key, reverse=not ascending)
        new_index = tuple(value.name for value in sorted_values)
        return MappedTable(values=sorted_values, index=new_index, columns=self.columns)

    def _get_empty_sequence(self, ):
        return (None,) * self.shape[0]

    def reindex(self, columns):
        values = [self.values.get(value, self._get_empty_sequence()) for value in columns]
        return MappedTable(values=values, index=self.index, columns=columns, axis=1)

    def to_list(self):
        return list(self.row_values)

    def to_dict(self):
        return dict(self.column_values)

    def where(self, func=None, **kwargs):
        def compare(x):
            return x[items] == values

        if func is None:
            items = list(kwargs.keys())
            values = tuple(kwargs.values())
            item_in_columns = [key in self.columns for key in kwargs.keys()]
            if not all(item_in_columns):
                raise KeyError
            func = compare
        new_keys = []
        new_values = []
        for value in self.row_values:
            if func(value):
                new_keys.append(value.name)
                new_values.append(value)

        return MappedTable(values=new_values, index=new_keys, columns=self.columns)

    def unique(self):
        return set(self.row_values)

    def isnone(self):
        new_values = [value.isnone() for value in self.row_values]
        return MappedTable(values=new_values, index=self.index, columns=self.columns, axis=0)

    def dropnone(self):
        new_values = [value for value in self.row_values if not any(value.isnone())]
        return MappedTable(values=new_values, index=self.index, columns=self.columns, axis=0)

    def fillnone(self, value):
        return MappedTable([column.fillnone(value) for column in self._column_values],
                           index=self.index, columns=self.columns, axis=1)

    def melt(self, id_vars, value_name=None, var_name=None):
        if value_name is None:
            value_name = 'value'
        if var_name is None:
            var_name = 'variable'

        if is_scalar(id_vars):
            assert id_vars in self.columns, \
                'id_vars should be in columns, expected {}, got {}'.format(self.columns, id_vars)
            id_vars = [id_vars]
        else:
            assert all([col in self.columns for col in id_vars]), \
                'id_vars should be in columns, expected {}, got {}'.format(self.columns, id_vars)
        column_to_melt = [col for col in self.columns if col not in id_vars]
        new_values = []
        for row in self.row_values:
            for col in column_to_melt:
                new_values.append((*row[id_vars], col, row[col]))

        new_columns = id_vars + [var_name, value_name]
        return MappedTable(new_values, columns=new_columns)

    def pivot(self, index: Union[str, List[str]], column, value, agg_func='mean'):
        unique_columns = self[column].unique()
        unique_index = self[index].unique()
        # define aggregation function
        if agg_func == 'mean':
            agg_func = mean_aggregate
        else:
            pass

        new_values = []
        for index_value in unique_index:
            if is_iterable(index):
                query = dict(zip(index, index_value))
                new_row = [*index_value]
            else:
                query = {index: index_value}
                new_row = [index_value]
            for column_value in unique_columns:
                # retrieve data corresponding to both index and column
                # TODO when multiple aggregation, return with a new column
                #  'agg' containing values for each aggregation function
                data = self.where(**query, **{column: column_value})[value]
                new_row.append(agg_func(data))
            new_values.append(new_row)

        return MappedTable(values=new_values, columns=[index, *unique_columns], axis=0)
