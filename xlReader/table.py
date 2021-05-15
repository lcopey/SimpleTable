from typing import Iterable, Optional, Union
from openpyxl import load_workbook
from .mapped_sequence import MappedSequence
from .utils import is_iterable


class Table:
    def __init__(self, rows: Iterable[Iterable], column_names: Optional[Iterable[str]] = None, ):
        self._column_names = column_names
        self._rows = MappedSequence(rows)

        # Build columns
        # new_columns = []
        # for n, name in enumerate(self._column_names):
        #     column = Column(n, name, self._rows, row_names=None)
        #     new_columns.append(column)
        # self._columns = MappedSequence(new_columns, self._column_names)
        self._columns = MappedSequence.from_transposed(rows, column_names)

    @classmethod
    def from_excel(cls, file_path, header: Optional[Union[int, Iterable[int]]] = 0, sheetname: Optional[str] = None):
        # Get workbook
        workbook = load_workbook(file_path, data_only=True)
        # Handle sheetname
        # By default, parse the first sheet in the excel
        if sheetname is None:
            sheetname = workbook.sheetnames[0]
        sheet = workbook.get_sheet_by_name(sheetname)

        # Handle header
        # 1. Instantiate column_names
        # 2. Instantiate rows
        if header is None:
            column_names = None
            rows = [row for n, row in enumerate(sheet.iter_rows(values_only=True))]
        elif is_iterable(header):
            column_names = list(
                zip(*[row for row in sheet.iter_rows(min_row=min(header), max_row=max(header) + 1, values_only=True)])
            )
            rows = [row for n, row in enumerate(sheet.iter_rows(values_only=True)) if n not in header]
        elif type(header) == int:
            column_names = [row for row in sheet.iter_rows(min_row=header, max_row=header + 1, values_only=True)][0]
            rows = [row for n, row in enumerate(sheet.iter_rows(values_only=True)) if n != header]
        else:
            raise ValueError

        return cls(rows=rows, column_names=column_names)

    def __len__(self):
        return self._rows.__len__()

    def __iter__(self):
        return self._rows.__iter__()

    def __getitem__(self, key):
        return self._rows.__getitem__(key)

    @property
    def rows(self):
        return self._rows

    @property
    def columns(self):
        return self._columns

    @property
    def column_names(self):
        return self._column_names
